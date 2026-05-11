import calendar
import json
import re
from io import BytesIO
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlencode

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components

try:
    from streamlit_js_eval import streamlit_js_eval
except Exception:
    streamlit_js_eval = None

try:
    from google_auth_oauthlib.flow import Flow
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build
except Exception:
    Flow = None
    Credentials = None
    build = None


# =========================
# App constants
# =========================

SCOPES = ["https://www.googleapis.com/auth/calendar.readonly"]
DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
OUTPUT_FILE = DATA_DIR / "availability_submissions.jsonl"

CALENDAR_LOCAL_STORAGE_KEY = "medstaff_calendar_events_by_date_v1"
EMPLOYEE_NAME_LOCAL_STORAGE_KEY = "medstaff_employee_name_v1"


def merge_calendar_events(existing_events: Dict[str, List[str]], new_events: Dict[str, List[str]]) -> Dict[str, List[str]]:
    merged_events = {}
    all_dates = set(existing_events.keys()) | set(new_events.keys())

    for date_key in all_dates:
        combined = []
        combined.extend(existing_events.get(date_key, []))
        combined.extend(new_events.get(date_key, []))
        merged_events[date_key] = list(dict.fromkeys(combined))

    return merged_events


def load_calendar_events_from_browser() -> Dict[str, List[str]]:
    if streamlit_js_eval is None:
        return {}

    try:
        raw_value = streamlit_js_eval(
            js_expressions=f"localStorage.getItem('{CALENDAR_LOCAL_STORAGE_KEY}')",
            key="load_calendar_events_local_storage",
        )
        if not raw_value:
            return {}

        payload = json.loads(raw_value)
        if not isinstance(payload, dict):
            return {}

        return {
            str(date_key): [str(item) for item in items]
            for date_key, items in payload.items()
            if isinstance(items, list)
        }
    except Exception:
        return {}


def save_calendar_events_to_browser(events_by_date: Dict[str, List[str]]) -> None:
    if streamlit_js_eval is None:
        return

    try:
        payload = json.dumps(events_by_date, ensure_ascii=False)
        streamlit_js_eval(
            js_expressions=f"localStorage.setItem('{CALENDAR_LOCAL_STORAGE_KEY}', {json.dumps(payload)});",
            key="save_calendar_events_local_storage",
        )
    except Exception:
        pass


def clear_calendar_events_from_browser() -> None:
    if streamlit_js_eval is None:
        return

    try:
        streamlit_js_eval(
            js_expressions=f"localStorage.removeItem('{CALENDAR_LOCAL_STORAGE_KEY}');",
            key="clear_calendar_events_local_storage",
        )
    except Exception:
        pass




st.set_page_config(
    page_title="מע׳ לתכנון תורנויות- פנימית ד׳",
    page_icon="🗓️",
    layout="wide",
)


def inject_global_rtl_css():
    st.markdown(
        """
        <style>
        html, body, .stApp, [data-testid="stAppViewContainer"] {
            direction: rtl !important;
            text-align: right !important;
            background: #fbfcfe;
        }

        .main .block-container {
            direction: rtl !important;
            text-align: right !important;
            padding-top: 2rem;
            max-width: 1540px;
        }

        h1, h2, h3, h4, h5, h6, p, label {
            text-align: right !important;
        }

        @media (min-width: 769px) {
            section[data-testid="stSidebar"] {
                right: 0 !important;
                left: auto !important;
                direction: rtl !important;
                text-align: right !important;
                border-left: 1px solid rgba(49, 51, 63, 0.12);
                border-right: none;
                background: #f6f8fb;
            }

            section[data-testid="stSidebar"] > div,
            div[data-testid="stSidebarContent"] {
                direction: rtl !important;
                text-align: right !important;
            }

            section[data-testid="stSidebar"] label,
            section[data-testid="stSidebar"] p,
            section[data-testid="stSidebar"] span,
            section[data-testid="stSidebar"] div {
                text-align: right !important;
            }

            .main .block-container {
                padding-right: 2rem;
                padding-left: 2rem;
            }
        }

        @media (max-width: 768px) {
            .main .block-container {
                padding: 1rem 0.75rem 5rem 0.75rem;
            }

            div[data-testid="stDataFrame"],
            div[data-testid="stDataEditor"] {
                overflow-x: auto;
            }

            button {
                width: 100%;
                min-height: 44px;
            }

            textarea, input {
                width: 100%;
                min-height: 42px;
            }

            .tool-hero, .month-nav-card, .choices-panel {
                padding: 1rem !important;
                border-radius: 18px !important;
            }

            .month-nav-grid {
                grid-template-columns: 1fr !important;
                gap: 0.75rem !important;
            }

            .summary-stat-grid {
                grid-template-columns: 1fr !important;
            }

            .month-center {
                order: -1;
            }

            .action-row {
                grid-template-columns: 1fr !important;
            }
        }

        table, textarea, input {
            direction: rtl !important;
            text-align: right !important;
        }

        div[data-testid="stHorizontalBlock"] {
            direction: rtl !important;
        }

        div[data-testid="stDataEditor"],
        div[data-testid="stDataFrame"] {
            direction: rtl !important;
            text-align: right !important;
        }

        div[data-testid="stDataEditor"] [role="grid"],
        div[data-testid="stDataFrame"] [role="grid"] {
            direction: rtl !important;
        }

        div[data-testid="stDataEditor"] [role="columnheader"],
        div[data-testid="stDataEditor"] [role="gridcell"],
        div[data-testid="stDataFrame"] [role="columnheader"],
        div[data-testid="stDataFrame"] [role="gridcell"] {
            text-align: right !important;
            justify-content: flex-end !important;
            direction: rtl !important;
        }

        button, [role="radiogroup"], [data-testid="stRadio"] {
            direction: rtl !important;
            text-align: right !important;
        }

        [data-testid="stRadio"] label {
            direction: rtl !important;
            text-align: right !important;
            justify-content: flex-start !important;
        }

        .tool-hero {
            background: radial-gradient(circle at top right, #ffffff 0%, #f8fbff 42%, #f4f8ff 100%);
            border: 1px solid #e8eef7;
            border-radius: 22px;
            padding: 1.4rem 1.6rem;
            margin-bottom: 1.05rem;
            box-shadow: 0 14px 32px rgba(31, 45, 61, 0.06);
        }

        .tool-hero h1 {
            margin: 0 0 0.35rem 0;
            font-size: 2rem;
            font-weight: 850;
            color: #202938;
        }

        .tool-hero p {
            margin: 0;
            color: #647084;
            font-size: 1rem;
        }

        .month-nav-card {
            background: rgba(255, 255, 255, 0.92);
            border: 1px solid #e9edf5;
            border-radius: 24px;
            padding: 1.35rem 1.55rem;
            box-shadow: 0 18px 44px rgba(31, 45, 61, 0.08);
            margin: 0.6rem 0 1.15rem 0;
        }

        .month-nav-grid {
            display: grid;
            grid-template-columns: minmax(210px, 1fr) minmax(340px, 1.2fr) minmax(260px, 1fr);
            align-items: center;
            gap: 1.15rem;
        }

        .month-center {
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 1rem;
            border-left: 1px solid #e0e6ef;
            border-right: 1px solid #e0e6ef;
            padding: 0.3rem 1rem;
        }

        .month-icon {
            width: 58px;
            height: 58px;
            border-radius: 999px;
            display: flex;
            align-items: center;
            justify-content: center;
            background: #e8f2ff;
            color: #2f7de1;
            font-size: 1.55rem;
            box-shadow: inset 0 0 0 1px rgba(47, 125, 225, 0.08);
        }

        .month-title-big {
            font-size: 2rem;
            font-weight: 900;
            color: #202938;
            line-height: 1;
            text-align: center !important;
        }

        .month-subtitle {
            margin-top: 0.55rem;
            display: inline-block;
            background: #e8f2ff;
            color: #2878d8;
            border-radius: 999px;
            padding: 0.32rem 0.9rem;
            font-size: 0.95rem;
            font-weight: 750;
        }

        .soft-note {
            background: #eaf4ff;
            border: 1px solid #d8e9ff;
            color: #1d5fae;
            border-radius: 14px;
            padding: 0.75rem 1rem;
            margin: 0.85rem 0;
            font-weight: 600;
        }

        .summary-stat-grid {
            display: grid;
            grid-template-columns: repeat(4, minmax(170px, 1fr));
            gap: 0.85rem;
            margin: 0.8rem 0 1.05rem 0;
        }

        .summary-card {
            border-radius: 18px;
            padding: 1rem 1.1rem;
            border: 1px solid #e8edf5;
            background: #fff;
            box-shadow: 0 10px 28px rgba(31, 45, 61, 0.045);
            min-height: 92px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }

        .summary-card.blocked {
            background: linear-gradient(135deg, #fff8f8 0%, #fff 100%);
            border-color: #ffd2d2;
        }

        .summary-card.vacation {
            background: linear-gradient(135deg, #f1fff6 0%, #fff 100%);
            border-color: #cdeed8;
        }

        .summary-card.total {
            background: linear-gradient(135deg, #f5faff 0%, #fff 100%);
            border-color: #d8e9ff;
        }

        .summary-card.days {
            background: linear-gradient(135deg, #fbfcff 0%, #fff 100%);
        }

        .summary-card-label {
            font-size: 0.9rem;
            color: #687385;
            font-weight: 700;
            margin-bottom: 0.35rem;
        }

        .summary-card-value {
            font-size: 1.45rem;
            color: #202938;
            font-weight: 900;
        }

        .summary-card-sub {
            color: #647084;
            font-size: 0.85rem;
        }

        .summary-icon {
            font-size: 1.7rem;
            opacity: 0.9;
        }

        .choices-panel {
            background: #ffffff;
            border: 1px solid #e9edf5;
            border-radius: 22px;
            padding: 1.15rem 1.25rem;
            box-shadow: 0 14px 34px rgba(31, 45, 61, 0.06);
            margin-top: 1rem;
        }

        .choices-panel-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 1rem;
            margin-bottom: 0.75rem;
        }

        .choices-panel-title {
            font-size: 1.35rem;
            font-weight: 900;
            color: #202938;
        }

        .choices-panel-subtitle {
            color: #647084;
            font-size: 0.94rem;
            margin-top: 0.2rem;
        }

        .action-row {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 0.85rem;
            align-items: center;
            margin-top: 0.85rem;
        }

        .success-strip {
            background: linear-gradient(90deg, #eafaf0 0%, #f5fff8 100%);
            border: 1px solid #cdeed8;
            border-radius: 16px;
            padding: 0.9rem 1rem;
            color: #16803b;
            font-weight: 750;
            display: flex;
            align-items: center;
            justify-content: space-between;
            margin-top: 0.9rem;
        }

        .sticky-actions {
            position: sticky;
            bottom: 0;
            z-index: 999;
            background: rgba(255, 255, 255, 0.97);
            backdrop-filter: blur(8px);
            border-top: 1px solid #e9edf3;
            padding: 0.85rem 0;
            margin-top: 1rem;
        }

        /* Strong RTL fixes for all tables and summary areas */
        div[data-testid="stDataFrame"] {
            direction: rtl !important;
        }

        div[data-testid="stDataFrame"] div[role="grid"],
        div[data-testid="stDataFrame"] div[role="row"],
        div[data-testid="stDataFrame"] div[role="columnheader"],
        div[data-testid="stDataFrame"] div[role="gridcell"] {
            direction: rtl !important;
            text-align: right !important;
        }

        div[data-testid="stDataFrame"] div[role="columnheader"],
        div[data-testid="stDataFrame"] div[role="gridcell"] {
            justify-content: flex-end !important;
        }

        .summary-table-rtl {
            direction: rtl !important;
            text-align: right !important;
            width: 100%;
        }

        .summary-table-rtl table {
            direction: rtl !important;
            text-align: right !important;
        }

        .summary-table-rtl th,
        .summary-table-rtl td {
            direction: rtl !important;
            text-align: right !important;
        }

        .copy-output textarea {
            direction: rtl !important;
            text-align: right !important;
            font-size: 1rem !important;
            line-height: 1.7 !important;
        }

        .bottom-actions-row {
            direction: rtl !important;
        }


        /* Tablet layout */
        @media (min-width: 769px) and (max-width: 1180px) {
            .main .block-container {
                padding: 1.2rem 1rem 4rem 1rem !important;
                max-width: 100% !important;
            }

            section[data-testid="stSidebar"] {
                width: 280px !important;
                min-width: 280px !important;
            }

            .tool-hero h1 {
                font-size: 1.55rem !important;
            }

            .month-nav-grid,
            .summary-stat-grid {
                grid-template-columns: 1fr 1fr !important;
            }
        }

        /* Phone layout */
        @media (max-width: 768px) {
            html, body, .stApp {
                overflow-x: hidden !important;
            }

            .main .block-container {
                padding: 0.75rem 0.6rem 5rem 0.6rem !important;
                max-width: 100% !important;
            }

            section[data-testid="stSidebar"] {
                width: 100% !important;
                min-width: 100% !important;
                max-width: 100% !important;
            }

            section[data-testid="stSidebar"] * {
                white-space: normal !important;
                overflow-wrap: anywhere !important;
                word-break: normal !important;
                line-height: 1.35 !important;
            }

            section[data-testid="stSidebar"] h1,
            section[data-testid="stSidebar"] h2,
            section[data-testid="stSidebar"] h3 {
                font-size: 1.05rem !important;
            }

            section[data-testid="stSidebar"] label,
            section[data-testid="stSidebar"] p,
            section[data-testid="stSidebar"] span {
                font-size: 0.92rem !important;
            }

            .tool-hero {
                padding: 0.9rem !important;
                margin-bottom: 0.75rem !important;
            }

            .tool-hero h1 {
                font-size: 1.35rem !important;
            }

            .tool-hero p {
                font-size: 0.9rem !important;
            }

            .month-nav-card,
            .choices-panel,
            .summary-box {
                padding: 0.85rem !important;
                border-radius: 16px !important;
                margin-top: 0.7rem !important;
            }

            .month-title-big {
                font-size: 1.35rem !important;
            }

            .month-icon {
                width: 44px !important;
                height: 44px !important;
                font-size: 1.15rem !important;
            }

            .month-center {
                border-left: none !important;
                border-right: none !important;
                padding: 0 !important;
            }

            .summary-stat-grid {
                grid-template-columns: 1fr 1fr !important;
                gap: 0.55rem !important;
            }

            .summary-card {
                min-height: 74px !important;
                padding: 0.75rem !important;
            }

            .summary-card-label,
            .summary-card-sub {
                font-size: 0.76rem !important;
            }

            .summary-card-value {
                font-size: 1.1rem !important;
            }

            .summary-icon {
                font-size: 1.25rem !important;
            }

            div[data-testid="stDataEditor"],
            div[data-testid="stDataFrame"] {
                max-width: 100vw !important;
                overflow-x: auto !important;
                -webkit-overflow-scrolling: touch !important;
            }

            div[data-testid="stDataEditor"] [role="grid"],
            div[data-testid="stDataFrame"] [role="grid"] {
                min-width: 780px !important;
            }

            button {
                min-height: 46px !important;
                font-size: 0.95rem !important;
                padding: 0.55rem 0.75rem !important;
            }

            textarea {
                min-height: 120px !important;
                font-size: 0.92rem !important;
            }

            input {
                min-height: 42px !important;
                font-size: 0.95rem !important;
            }

            div[data-baseweb="select"] {
                max-width: 100% !important;
            }

            div[data-baseweb="tag"] {
                max-width: 100% !important;
                margin: 0.15rem !important;
            }

            div[data-baseweb="tag"] span {
                overflow: hidden !important;
                text-overflow: ellipsis !important;
                white-space: nowrap !important;
                max-width: 230px !important;
            }

            .bottom-actions-row,
            .action-row {
                grid-template-columns: 1fr !important;
            }
        }


        /* Final mobile sidebar override: do not force right sidebar on phones */
        @media (max-width: 768px) {
            section[data-testid="stSidebar"] {
                position: fixed !important;
                left: 0 !important;
                right: auto !important;
                width: min(92vw, 360px) !important;
                min-width: min(92vw, 360px) !important;
                max-width: min(92vw, 360px) !important;
                direction: rtl !important;
                text-align: right !important;
                overflow-x: hidden !important;
            }

            div[data-testid="stSidebarContent"] {
                width: 100% !important;
                max-width: 100% !important;
                overflow-x: hidden !important;
                padding-left: 0.75rem !important;
                padding-right: 0.75rem !important;
            }

            section[data-testid="stSidebar"] div,
            section[data-testid="stSidebar"] label,
            section[data-testid="stSidebar"] p,
            section[data-testid="stSidebar"] span {
                max-width: 100% !important;
                white-space: normal !important;
                overflow-wrap: break-word !important;
                word-break: normal !important;
            }

            [data-testid="collapsedControl"] {
                left: 0.6rem !important;
                right: auto !important;
            }
        }

        @media (max-width: 900px) and (orientation: landscape) {
            section[data-testid="stSidebar"] {
                width: min(86vw, 420px) !important;
                max-width: min(86vw, 420px) !important;
            }
        }

        </style>
        """,
        unsafe_allow_html=True,
    )

def load_employee_name_from_browser() -> str:
    if streamlit_js_eval is None:
        return ""

    try:
        value = streamlit_js_eval(
            js_expressions=f"localStorage.getItem('{EMPLOYEE_NAME_LOCAL_STORAGE_KEY}')",
            key="load_employee_name_local_storage",
        )
        return str(value or "")
    except Exception:
        return ""


def save_employee_name_to_browser(employee_name: str) -> None:
    if streamlit_js_eval is None:
        return

    try:
        streamlit_js_eval(
            js_expressions=f"localStorage.setItem('{EMPLOYEE_NAME_LOCAL_STORAGE_KEY}', {json.dumps(employee_name or '', ensure_ascii=False)});",
            key="save_employee_name_local_storage",
        )
    except Exception:
        pass


# =========================
# Date helpers
# =========================

def add_months(year: int, month: int, delta: int) -> Tuple[int, int]:
    """Return year, month after adding delta months."""
    zero_based = (year * 12 + (month - 1)) + delta
    return zero_based // 12, zero_based % 12 + 1


def default_next_month() -> Tuple[int, int]:
    today = date.today()
    return add_months(today.year, today.month, 1)


def default_previous_month() -> Tuple[int, int]:
    today = date.today()
    return add_months(today.year, today.month, -1)


def default_current_month() -> Tuple[int, int]:
    today = date.today()
    return today.year, today.month


def month_dates(year: int, month: int) -> List[date]:
    last_day = calendar.monthrange(year, month)[1]
    return [date(year, month, day) for day in range(1, last_day + 1)]


def iso_month_range(year: int, month: int) -> Tuple[str, str]:
    days = month_dates(year, month)
    start_dt = datetime.combine(days[0], time.min).astimezone()
    end_dt = datetime.combine(days[-1] + timedelta(days=1), time.min).astimezone()
    return start_dt.isoformat(), end_dt.isoformat()


def hebrew_weekday(d: date) -> str:
    names = {
        0: "שני",
        1: "שלישי",
        2: "רביעי",
        3: "חמישי",
        4: "שישי",
        5: "שבת",
        6: "ראשון",
    }
    return names[d.weekday()]


def month_title(year: int, month: int) -> str:
    hebrew_months = {
        1: "ינואר",
        2: "פברואר",
        3: "מרץ",
        4: "אפריל",
        5: "מאי",
        6: "יוני",
        7: "יולי",
        8: "אוגוסט",
        9: "ספטמבר",
        10: "אוקטובר",
        11: "נובמבר",
        12: "דצמבר",
    }
    return f"{hebrew_months[month]} {year}"


# =========================
# Holiday provider
# MVP: display-only Jewish/Israeli holidays.
# This is intentionally separate from the scheduling engine.
# =========================

@st.cache_data(ttl=60 * 60 * 24)
def fetch_jewish_holidays_from_hebcal(year: int) -> Dict[str, List[str]]:
    """
    Fetch Jewish holidays for Israel from Hebcal.
    Display-only; user decides what becomes unavailable/vacation.
    """
    url = "https://www.hebcal.com/hebcal"
    params = {
        "v": "1",
        "cfg": "json",
        "maj": "on",
        "min": "on",
        "mod": "on",
        "nx": "on",
        "year": year,
        "month": "x",
        "geo": "none",
        "c": "off",
        "lg": "he",
        "i": "on",
    }

    try:
        res = requests.get(url, params=params, timeout=10)
        res.raise_for_status()
        payload = res.json()
    except Exception:
        return {}

    holidays: Dict[str, List[str]] = {}
    for item in payload.get("items", []):
        category = item.get("category", "")
        if category not in {"holiday", "hebdate", "roshchodesh", "fast"}:
            continue
        iso = item.get("date")
        title = item.get("title")
        if not iso or not title:
            continue
        # Normalize date part only.
        iso_date = iso[:10]
        holidays.setdefault(iso_date, []).append(title)
    return holidays


def holiday_map_for_month(year: int, month: int) -> Dict[str, str]:
    # For January/December month views, fetch current year only is enough for visible dates.
    raw = fetch_jewish_holidays_from_hebcal(year)
    out = {}
    for d in month_dates(year, month):
        values = raw.get(d.isoformat(), [])
        if d.weekday() == 5:
            values = ["שבת"] + values
        out[d.isoformat()] = " | ".join(values)
    return out


# =========================
# Calendar integrations
# =========================

def parse_google_secret_client_config() -> Optional[dict]:
    """
    Expected Streamlit secrets format:

    [google_oauth]
    client_id = "..."
    client_secret = "..."
    redirect_uri = "https://your-app.streamlit.app"
    """
    try:
        cfg = st.secrets["google_oauth"]
        return {
            "web": {
                "client_id": cfg["client_id"],
                "client_secret": cfg["client_secret"],
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [cfg["redirect_uri"]],
            }
        }
    except Exception:
        return None


def google_is_configured() -> bool:
    return Flow is not None and build is not None and parse_google_secret_client_config() is not None



def _new_google_flow() -> Optional[Flow]:
    client_config = parse_google_secret_client_config()
    if not client_config or Flow is None:
        return None

    redirect_uri = st.secrets["google_oauth"]["redirect_uri"]

    # Streamlit Cloud may return from Google in a refreshed session.
    # Use a confidential Web OAuth client with client_secret, without PKCE/code_verifier.
    return Flow.from_client_config(
        client_config,
        scopes=SCOPES,
        redirect_uri=redirect_uri,
        autogenerate_code_verifier=False,
    )


def start_google_oauth() -> Optional[str]:
    if Flow is None:
        return None

    flow = _new_google_flow()
    if flow is None:
        return None

    auth_url, state = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent",
    )
    st.session_state["google_oauth_state"] = state
    return auth_url


def finish_google_oauth(code: str) -> bool:
    if Flow is None:
        return False

    flow = _new_google_flow()
    if flow is None:
        return False

    flow.fetch_token(code=code)
    creds = flow.credentials

    st.session_state["google_credentials"] = {
        "token": creds.token,
        "refresh_token": creds.refresh_token,
        "token_uri": creds.token_uri,
        "client_id": creds.client_id,
        "client_secret": creds.client_secret,
        "scopes": creds.scopes,
    }

    st.session_state.pop("google_oauth_state", None)

    try:
        st.query_params.clear()
    except Exception:
        pass

    return True


def list_google_calendars() -> List[dict]:
    if "google_credentials" not in st.session_state or Credentials is None or build is None:
        return []

    creds = Credentials(**st.session_state["google_credentials"])
    service = build("calendar", "v3", credentials=creds)
    result = service.calendarList().list().execute()
    return result.get("items", [])



def _format_google_event_time(event: dict) -> str:
    start = event.get("start", {})
    end = event.get("end", {})

    if start.get("date"):
        return "כל היום"

    start_raw = start.get("dateTime")
    end_raw = end.get("dateTime")
    if not start_raw:
        return ""

    def parse_time(raw: str) -> str:
        try:
            dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
            return dt.strftime("%H:%M")
        except Exception:
            return raw[11:16] if len(raw) >= 16 else ""

    start_time = parse_time(start_raw)
    end_time = parse_time(end_raw) if end_raw else ""
    return f"{start_time}-{end_time}" if start_time and end_time else start_time


def read_google_events(calendar_ids: List[str], year: int, month: int) -> Dict[str, List[str]]:
    if "google_credentials" not in st.session_state or Credentials is None or build is None:
        return {}

    time_min, time_max = iso_month_range(year, month)
    creds = Credentials(**st.session_state["google_credentials"])
    service = build("calendar", "v3", credentials=creds)

    calendar_names = {}
    try:
        for cal in service.calendarList().list().execute().get("items", []):
            calendar_names[cal.get("id")] = cal.get("summary", cal.get("id", "יומן"))
    except Exception:
        pass

    events_by_date: Dict[str, List[str]] = {}
    for cal_id in calendar_ids:
        cal_name = calendar_names.get(cal_id, cal_id)
        result = service.events().list(
            calendarId=cal_id,
            timeMin=time_min,
            timeMax=time_max,
            singleEvents=True,
            orderBy="startTime",
            maxResults=2500,
        ).execute()

        for event in result.get("items", []):
            start = event.get("start", {})
            raw_start = start.get("date") or start.get("dateTime")
            if not raw_start:
                continue

            iso_date = raw_start[:10]
            title = event.get("summary", "אירוע ללא כותרת")
            time_text = _format_google_event_time(event)
            label = f"{time_text} [{cal_name}] {title}" if time_text else f"[{cal_name}] {title}"
            events_by_date.setdefault(iso_date, []).append(label)

    return events_by_date


def events_to_cell(events_by_date: Dict[str, List[str]], d: date) -> str:
    events = events_by_date.get(d.isoformat(), [])
    if not events:
        return ""
    # Keep the table readable.
    return " | ".join(events[:5]) + (" ..." if len(events) > 5 else "")


# =========================
# Availability table
# =========================

def build_month_dataframe(year: int, month: int, events_by_date: Dict[str, List[str]]) -> pd.DataFrame:
    holidays = holiday_map_for_month(year, month)
    rows = []

    for d in month_dates(year, month):
        rows.append({
            "date": d.isoformat(),
            "הערה": "",
            "יום חופש": False,
            "חסום לתורנות": False,
            "אירועים מהיומן": events_to_cell(events_by_date, d),
            "חגים / שבתות": holidays.get(d.isoformat(), ""),
            "יום": hebrew_weekday(d),
            "תאריך": d.strftime("%d/%m/%Y"),
        })

    return pd.DataFrame(rows)



def summarize_submission(df: pd.DataFrame, person_id: str, year: int, month: int) -> List[dict]:
    constraints = []

    for _, row in df.iterrows():
        if bool(row.get("חסום לתורנות")):
            constraints.append({
                "person_id": person_id,
                "date": row["date"],
                "day_in_month": int(str(row["date"])[8:10]),
                "type": "unavailable_for_shift",
                "strength": "hard",
                "source": "manual",
                "note": str(row.get("הערה", "") or ""),
            })

        if bool(row.get("יום חופש")):
            constraints.append({
                "person_id": person_id,
                "date": row["date"],
                "day_in_month": int(str(row["date"])[8:10]),
                "type": "vacation_request",
                "strength": "hard",
                "source": "manual",
                "note": str(row.get("הערה", "") or ""),
            })

    return constraints


def build_copyable_submission_text(df: pd.DataFrame, employee_name: str) -> str:
    blocked_days = []
    vacation_days = []

    for _, row in df.iterrows():
        day_num = int(str(row["date"])[8:10])
        if bool(row.get("חסום לתורנות")):
            blocked_days.append(day_num)
        if bool(row.get("יום חופש")):
            vacation_days.append(day_num)

    blocked = ",".join(str(x) for x in blocked_days)
    vacations = ",".join(str(x) for x in vacation_days)

    return f"{employee_name}\nחסימות- {blocked}\nחופשים- {vacations}"




def get_selected_days(df: pd.DataFrame) -> Tuple[List[int], List[int]]:
    blocked_days = []
    vacation_days = []

    for _, row in df.iterrows():
        day_num = int(str(row["date"])[8:10])
        if bool(row.get("חסום לתורנות")):
            blocked_days.append(day_num)
        if bool(row.get("יום חופש")):
            vacation_days.append(day_num)

    return blocked_days, vacation_days


def render_stat_cards(blocked_days: List[int], vacation_days: List[int], notes_count: int):
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(
            f"""
            <div class="stat-card">
                <div class="stat-label">חסימות לתורנות</div>
                <div class="stat-value">{len(blocked_days)}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c2:
        st.markdown(
            f"""
            <div class="stat-card">
                <div class="stat-label">ימי חופש</div>
                <div class="stat-value">{len(vacation_days)}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c3:
        st.markdown(
            f"""
            <div class="stat-card">
                <div class="stat-label">הערות</div>
                <div class="stat-value">{notes_count}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )


def style_availability_preview(df: pd.DataFrame):
    styles = pd.DataFrame("", index=df.index, columns=df.columns)

    for idx, row in df.iterrows():
        is_weekend_or_holiday = row.get("יום") in ["שישי", "שבת"] or bool(str(row.get("חגים / שבתות", "")).strip())
        has_block = bool(row.get("חסום לתורנות"))
        has_vacation = bool(row.get("יום חופש"))

        for col in df.columns:
            if is_weekend_or_holiday and col in ["תאריך", "יום", "חגים / שבתות"]:
                styles.loc[idx, col] = "background-color: #fff3bf; font-weight: 700;"
            elif has_vacation and col == "יום חופש":
                styles.loc[idx, col] = "background-color: #d3f9d8;"
            elif has_block and col == "חסום לתורנות":
                styles.loc[idx, col] = "background-color: #ffe3e3;"
            elif idx % 2 == 1:
                styles.loc[idx, col] = "background-color: #fafafa;"

    return df.style.apply(lambda _: styles, axis=None)


def availability_table_to_xlsx_bytes(df: pd.DataFrame, employee_name: str, year: int, month: int) -> bytes:
    export_cols = ["תאריך", "יום", "חגים / שבתות", "אירועים מהיומן", "חסום לתורנות", "יום חופש", "הערה"]
    export_df = df[export_cols].copy()

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="זמינות", index=False, startrow=3)
        wb = writer.book
        ws = wb["זמינות"]
        ws.sheet_view.rightToLeft = True

        # Title area
        ws["A1"] = "מע׳ לתכנון תורנויות- פנימית ד׳"
        ws["A2"] = f"שם העובד/ת: {employee_name}"
        ws["D2"] = f"חודש: {month_title(year, month)}"

        title_fill = PatternFill("solid", fgColor="EAF4FF")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        weekend_fill = PatternFill("solid", fgColor="FFF3BF")
        blocked_fill = PatternFill("solid", fgColor="FFE3E3")
        vacation_fill = PatternFill("solid", fgColor="D3F9D8")
        thin = Side(style="thin", color="D9DDE5")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = title_fill
        ws["A1"].font = Font(bold=True, size=15)
        ws["A2"].font = Font(bold=True)
        ws["D2"].font = Font(bold=True)

        header_row = 4
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        max_row = ws.max_row
        max_col = ws.max_column
        for row_idx in range(header_row + 1, max_row + 1):
            weekday = ws.cell(row=row_idx, column=2).value
            holiday_value = ws.cell(row=row_idx, column=3).value
            blocked_value = ws.cell(row=row_idx, column=5).value
            vacation_value = ws.cell(row=row_idx, column=6).value

            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

                if row_idx % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor="FAFAFA")

            if weekday in ["שישי", "שבת"] or bool(holiday_value):
                for col_idx in [1, 2, 3]:
                    ws.cell(row=row_idx, column=col_idx).fill = weekend_fill
                    ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)

            if blocked_value is True:
                ws.cell(row=row_idx, column=5).fill = blocked_fill

            if vacation_value is True:
                ws.cell(row=row_idx, column=6).fill = vacation_fill

        widths = {
            1: 14,
            2: 12,
            3: 18,
            4: 28,
            5: 16,
            6: 14,
            7: 32,
        }
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A5"

    return buffer.getvalue()


def render_mobile_cards(df: pd.DataFrame):
    st.markdown('<div class="mobile-help">במובייל מומלץ לגלול את הטבלה לרוחב. הכרטיסים הבאים נותנים תצוגת קריאה מהירה של הימים.</div>', unsafe_allow_html=True)
    preview_cols = ["תאריך", "יום", "חגים / שבתות", "חסום לתורנות", "יום חופש", "הערה"]
    st.dataframe(df[preview_cols], hide_index=True, use_container_width=True)

def persist_submission(payload: dict) -> None:
    with OUTPUT_FILE.open("a", encoding="utf-8") as f:
        f.write(json.dumps(payload, ensure_ascii=False) + "\n")


# =========================
# UI
# =========================

def init_state():
    if "selected_year" not in st.session_state or "selected_month" not in st.session_state:
        y, m = default_next_month()
        st.session_state["selected_year"] = y
        st.session_state["selected_month"] = m

    if "events_by_date" not in st.session_state:
        st.session_state["events_by_date"] = load_calendar_events_from_browser()

    if "employee_name" not in st.session_state:
        st.session_state["employee_name"] = load_employee_name_from_browser()


def move_month(delta: int):
    y, m = add_months(st.session_state["selected_year"], st.session_state["selected_month"], delta)
    st.session_state["selected_year"] = y
    st.session_state["selected_month"] = m



def render_copy_button(text_to_copy: str, button_label: str = "העתק"):
    escaped = json.dumps(text_to_copy, ensure_ascii=False)
    components.html(
        f"""
        <div dir="rtl" style="text-align:right; font-family:Arial, sans-serif;">
            <button
                onclick='navigator.clipboard.writeText({escaped}).then(() => {{
                    const el = document.getElementById("copy-status");
                    el.innerText = "הועתק";
                    setTimeout(() => el.innerText = "", 1800);
                }})'
                style="
                    background:linear-gradient(135deg, #3b82f6 0%, #2563eb 100%);
                    color:white;
                    border:none;
                    border-radius:8px;
                    padding:0.55rem 1.1rem;
                    cursor:pointer;
                    font-size:1rem;
                    width:100%;
                    max-width:220px;
                "
            >
                {button_label}
            </button>
            <span id="copy-status" style="margin-right:12px; color:#0a7f28; font-weight:bold;"></span>
        </div>
        """,
        height=55,
    )



def render_month_nav_card(year: int, month: int):
    days_count = calendar.monthrange(year, month)[1]
    st.markdown(
        f"""
        <div class="month-nav-card">
            <div class="month-nav-grid">
                <div></div>
                <div class="month-center">
                    <div class="month-icon">📅</div>
                    <div>
                        <div class="month-title-big">{month_title(year, month)}</div>
                        <div class="month-subtitle">{days_count} ימים בחודש</div>
                    </div>
                </div>
                <div></div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_summary_cards_for_choices(blocked_days: List[int], vacation_days: List[int], notes_count: int, year: int, month: int):
    total = len(blocked_days) + len(vacation_days)
    days_count = calendar.monthrange(year, month)[1]
    st.markdown(
        f"""
        <div class="summary-stat-grid">
            <div class="summary-card blocked">
                <div>
                    <div class="summary-card-label">חסומים לתורנות</div>
                    <div class="summary-card-value">{len(blocked_days)}</div>
                    <div class="summary-card-sub">ימים</div>
                </div>
                <div class="summary-icon">🚫</div>
            </div>
            <div class="summary-card vacation">
                <div>
                    <div class="summary-card-label">ימי חופש</div>
                    <div class="summary-card-value">{len(vacation_days)}</div>
                    <div class="summary-card-sub">ימים</div>
                </div>
                <div class="summary-icon">🌴</div>
            </div>
            <div class="summary-card total">
                <div>
                    <div class="summary-card-label">סה״כ בחירות</div>
                    <div class="summary-card-value">{total}</div>
                    <div class="summary-card-sub">שורות</div>
                </div>
                <div class="summary-icon">☷</div>
            </div>
            <div class="summary-card days">
                <div>
                    <div class="summary-card-label">ימי החודש</div>
                    <div class="summary-card-value">{days_count}</div>
                    <div class="summary-card-sub">ימים</div>
                </div>
                <div class="summary-icon">🗓️</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_shift_planning():
    init_state()

    st.markdown("""<div class="tool-hero"><h1>תכנון תורנויות</h1><p>בחר ימים בהם אינך זמין לתורנויות או מעוניין בחופש. בסיום ניתן להעתיק פלט או להוריד XLSX.</p></div>""", unsafe_allow_html=True)

    with st.sidebar:
        st.subheader("הגדרות תכנון")
        employee_name = st.text_input(
            "שם העובד/ת",
            value=st.session_state.get("employee_name", ""),
            placeholder="הקלד/י שם מלא",
            key="employee_name_input",
        )
        st.session_state["employee_name"] = employee_name
        save_employee_name_to_browser(employee_name)
        person_id = employee_name

        st.divider()
        st.subheader("חיבור יומנים אישיים")
        st.caption("Google Calendar, קריאה בלבד. ניתן לבחור כמה יומנים, להסיר יומנים לא רצויים ואז לטעון אירועים.")
        if streamlit_js_eval is None:
            st.warning("שמירה מקומית בדפדפן אינה זמינה כי חסרה הספרייה streamlit-js-eval. האירועים יישמרו רק עד רענון הדף.")

        if not google_is_configured():
            st.warning("Google Calendar לא מוגדר עדיין.")
            st.markdown(
                """
                כדי שהחיבור יעבוד צריך:
                1. לפתוח פרויקט ב-Google Cloud.
                2. להפעיל Google Calendar API.
                3. ליצור OAuth Client מסוג Web application.
                4. להגדיר Authorized redirect URI זהה לכתובת האפליקציה ב-Streamlit Cloud.
                5. להוסיף את הערכים ב-Secrets של Streamlit Cloud.
                """
            )
            st.code(
                """[google_oauth]
client_id = "YOUR_CLIENT_ID"
client_secret = "YOUR_CLIENT_SECRET"
redirect_uri = "https://YOUR_APP.streamlit.app"
""",
                language="toml",
            )
        else:
            query_params = st.query_params
            code = query_params.get("code")

            if code and "google_credentials" not in st.session_state:
                try:
                    if finish_google_oauth(code):
                        st.session_state["events_by_date"] = merge_calendar_events(
                            load_calendar_events_from_browser(),
                            st.session_state.get("events_by_date", {}),
                        )
                        st.session_state["employee_name"] = load_employee_name_from_browser() or st.session_state.get("employee_name", "")
                        st.success("החיבור ליומן הושלם.")
                except Exception as e:
                    st.error(f"שגיאה בהשלמת החיבור: {e}")
                    st.info("אם הבעיה חוזרת: ודא שה-Secrets מעודכנים, שה-redirect_uri זהה ב-Google Cloud וב-Streamlit, ונסה לפתוח בדפדפן פרטי.")

            if "google_credentials" not in st.session_state:
                auth_url = start_google_oauth()
                if auth_url:
                    st.link_button("התחבר ל-Google Calendar", auth_url)
            else:
                st.success("מחובר ל-Google Calendar")
                try:
                    calendars = list_google_calendars()
                    options = {
                        f"{c.get('summary', c.get('id'))} ({c.get('id')})": c["id"]
                        for c in calendars
                    }

                    all_labels = list(options.keys())
                    selected_labels = st.multiselect(
                        "בחר יומנים לקריאה",
                        all_labels,
                        default=all_labels,
                        help="כל היומנים מסומנים כברירת מחדל. ניתן להסיר יומנים לא רצויים לפני הטעינה.",
                    )
                    selected_calendar_ids = [options[label] for label in selected_labels]

                    st.caption("האירועים יוצגו עם שעה ושם היומן. טעינת יומן נוסף תתווסף לאירועים שכבר נטענו. הנתונים נשמרים במכשיר הנוכחי בלבד.")

                    if st.button("טען אירועים מכל היומנים שנבחרו"):
                        if not selected_calendar_ids:
                            st.warning("יש לבחור לפחות יומן אחד.")
                        else:
                            save_employee_name_to_browser(st.session_state.get("employee_name", ""))
                            new_events = read_google_events(
                                selected_calendar_ids,
                                st.session_state["selected_year"],
                                st.session_state["selected_month"],
                            )

                            browser_events = load_calendar_events_from_browser()
                            existing_events = merge_calendar_events(
                                browser_events,
                                st.session_state.get("events_by_date", {}),
                            )
                            merged_events = merge_calendar_events(existing_events, new_events)

                            st.session_state["events_by_date"] = merged_events
                            save_calendar_events_to_browser(merged_events)
                            st.session_state["selected_calendar_count"] = len(selected_calendar_ids)

                            added_count = sum(len(v) for v in new_events.values())
                            total_count = sum(len(v) for v in merged_events.values())
                            st.success(f"נוספו {added_count} אירועים. סה״כ מוצגים כעת {total_count} אירועים.")

                    if st.button("נקה אירועים מהמכשיר הזה"):
                        st.session_state["events_by_date"] = {}
                        clear_calendar_events_from_browser()
                        st.success("האירועים נוקו מהמכשיר הנוכחי בלבד.")

                    if st.button("נתק חיבור ליומן"):
                        save_calendar_events_to_browser(st.session_state.get("events_by_date", {}))
                        st.session_state.pop("google_credentials", None)
                        st.info("החיבור נותק. האירועים שכבר נטענו נשארו במכשיר הזה ולא יימחקו.")
                        st.rerun()
                except Exception as e:
                    st.error(f"שגיאה בקריאת יומנים: {e}")


    # Month navigation
    nav_right, nav_center, nav_left = st.columns([1, 2.1, 1])
    with nav_right:
        if st.button("חודש קודם ←", use_container_width=True):
            move_month(-1)
            st.rerun()
    with nav_center:
        render_month_nav_card(st.session_state["selected_year"], st.session_state["selected_month"])
    with nav_left:
        nav_a, nav_b = st.columns(2)
        with nav_a:
            if st.button("→ חודש הבא", use_container_width=True):
                move_month(1)
                st.rerun()
        with nav_b:
            if st.button("📅 חזור לחודש הבא", use_container_width=True):
                y, m = default_next_month()
                st.session_state["selected_year"] = y
                st.session_state["selected_month"] = m
                st.rerun()

    y = st.session_state["selected_year"]
    m = st.session_state["selected_month"]

    df = build_month_dataframe(y, m, st.session_state["events_by_date"])

    st.info("ברירת המחדל היא החודש הבא. ניתן לנוע קדימה ואחורה ללא מגבלה כדי לתכנן חצי שנה ומעלה.")

    edited_df = st.data_editor(
        df,
        key=f"availability_editor_{y}_{m}",
        hide_index=True,
        use_container_width=True,
        column_order=["הערה", "יום חופש", "חסום לתורנות", "אירועים מהיומן", "חגים / שבתות", "יום", "תאריך"],
        disabled=["date", "תאריך", "יום", "חגים / שבתות", "אירועים מהיומן"],
        column_config={
            "date": None,
            "חסום לתורנות": st.column_config.CheckboxColumn("חסום לתורנות"),
            "יום חופש": st.column_config.CheckboxColumn("יום חופש"),
            "הערה": st.column_config.TextColumn("הערה"),
        },
    )

    blocked_days_live, vacation_days_live = get_selected_days(edited_df)
    notes_count_live = int(edited_df["הערה"].astype(str).str.strip().replace("nan", "").ne("").sum())
    render_summary_cards_for_choices(blocked_days_live, vacation_days_live, notes_count_live, y, m)

    st.divider()

    st.markdown('<div class="choices-panel">', unsafe_allow_html=True)
    col_preview, col_finish = st.columns([3, 1])
    constraints = summarize_submission(edited_df, person_id, y, m)
    submission_text = build_copyable_submission_text(edited_df, employee_name)

    with col_finish:
        submitted = st.button("שמור והפק פלט", type="primary", use_container_width=True)

    with col_preview:
        st.markdown("""<div class="choices-panel-header"><div><div class="choices-panel-title">סיכום בחירות</div><div class="choices-panel-subtitle">רשימת כל הבחירות שבוצעו במהלך החודש</div></div></div>""", unsafe_allow_html=True)
        if constraints:
            summary_df = pd.DataFrame(constraints)
            visible_columns = ["date", "day_in_month", "type", "note"]
            summary_df = summary_df[[col for col in visible_columns if col in summary_df.columns]]
            summary_df = summary_df.rename(columns={
                "date": "תאריך",
                "day_in_month": "יום בחודש",
                "type": "סוג",
                "note": "הערה",
            })
            summary_df["סוג"] = summary_df["סוג"].replace({
                "unavailable_for_shift": "חסום לתורנות",
                "vacation_request": "יום חופש",
            })
            # Streamlit's grid is visually LTR even under RTL CSS.
            # Reversing the technical order makes "תאריך" appear as the rightmost column.
            summary_display_order = ["הערה", "סוג", "יום בחודש", "תאריך"]
            summary_df = summary_df[[col for col in summary_display_order if col in summary_df.columns]]
            st.markdown('<div class="summary-table-rtl">', unsafe_allow_html=True)
            st.dataframe(summary_df, hide_index=True, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.caption("לא נבחרו חסימות או חופשות בחודש זה.")
    st.markdown("</div>", unsafe_allow_html=True)

    if submitted:
        payload = {
            "submitted_at": datetime.now(timezone.utc).isoformat(),
            "display_month": f"{y}-{m:02d}",
            "person_id": person_id,
            "employee_name": employee_name,
            "copyable_text": submission_text,
            "constraints": constraints,
        }
        persist_submission(payload)
        st.markdown(f"""<div class="success-strip"><span>✅ הפלט מוכן עבור {month_title(y, m)}.</span><span></span></div>""", unsafe_allow_html=True)

        st.markdown('<div class="choices-panel">', unsafe_allow_html=True)
        st.markdown("""<div class="choices-panel-title">פלט להעתקה ולהורדה</div>""", unsafe_allow_html=True)
        st.caption("העתק את הטקסט הבא ושלח אותו לריכוז. המבנה כולל שם, חסימות וחופשים לפי מספרי הימים בחודש.")
        st.markdown('<div class="copy-output">', unsafe_allow_html=True)
        st.text_area(
            "טקסט להעתקה",
            value=submission_text,
            height=120,
            label_visibility="collapsed",
        )
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="bottom-actions-row">', unsafe_allow_html=True)
        xlsx_col, copy_col = st.columns(2)
        with copy_col:
            render_copy_button(submission_text, "העתק")
        with xlsx_col:
            xlsx_bytes = availability_table_to_xlsx_bytes(edited_df, employee_name, y, m)
            st.download_button(
                "הורד XLSX",
                data=xlsx_bytes,
                file_name=f"availability_{employee_name}_{y}_{m:02d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        st.markdown('</div>', unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)



def parse_worker_blocks(raw_text: str) -> pd.DataFrame:
    """
    Expected block per worker:
    יאיר
    חסימות- 1,2,3
    חופשים- 1,11

    Multiple workers can be pasted into the same text area.
    """
    lines = [line.strip() for line in raw_text.splitlines() if line.strip()]
    records = []
    i = 0

    while i < len(lines):
        name = lines[i].strip()
        blocked_line = lines[i + 1] if i + 1 < len(lines) else ""
        vacation_line = lines[i + 2] if i + 2 < len(lines) else ""

        def extract_days(line: str) -> List[int]:
            if "-" in line:
                line = line.split("-", 1)[1]
            parts = re.split(r"[,;\\s]+", line.strip())
            days = []
            for part in parts:
                if not part:
                    continue
                try:
                    day = int(part)
                    if 1 <= day <= 31:
                        days.append(day)
                except ValueError:
                    pass
            return sorted(set(days))

        if name and not name.startswith("חסימות") and not name.startswith("חופשים"):
            records.append({
                "שם עובד/ת": name,
                "חסימות": ",".join(str(d) for d in extract_days(blocked_line)),
                "חופשים": ",".join(str(d) for d in extract_days(vacation_line)),
            })

        i += 3

    parsed_df = pd.DataFrame(records)
    if not parsed_df.empty:
        parsed_df["מספר חסימות"] = parsed_df["חסימות"].apply(lambda x: len([v for v in str(x).split(",") if v]))
        parsed_df["מספר חופשים"] = parsed_df["חופשים"].apply(lambda x: len([v for v in str(x).split(",") if v]))

    return parsed_df


def _days_from_csv(value: str) -> set:
    out = set()
    for part in str(value or "").split(","):
        part = part.strip()
        if part.isdigit():
            out.add(int(part))
    return out


def build_schedule_layout_table(parsed_df: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    previous_year, previous_month = add_months(year, month, -1)
    previous_last_day = calendar.monthrange(previous_year, previous_month)[1]
    current_last_day = calendar.monthrange(year, month)[1]

    row_dates = [
        date(previous_year, previous_month, previous_last_day - 1),
        date(previous_year, previous_month, previous_last_day),
    ] + [date(year, month, day) for day in range(1, current_last_day + 1)]

    employee_names = parsed_df["שם עובד/ת"].tolist() if not parsed_df.empty else []
    blocked_by_employee = {row["שם עובד/ת"]: _days_from_csv(row["חסימות"]) for _, row in parsed_df.iterrows()}
    vacation_by_employee = {row["שם עובד/ת"]: _days_from_csv(row["חופשים"]) for _, row in parsed_df.iterrows()}

    rows = []
    for d in row_dates:
        is_current_month = d.year == year and d.month == month
        out = {
            "יום בחודש": d.day,
            "יום בשבוע": hebrew_weekday(d),
            "מחלקה": "",
            "מיון": "",
            "שישי בוקר": "",
        }

        for employee in employee_names:
            blocked = is_current_month and d.day in blocked_by_employee.get(employee, set())
            vacation = is_current_month and d.day in vacation_by_employee.get(employee, set())
            out[employee] = "V" if (blocked or vacation) else ""

        rows.append(out)

    return pd.DataFrame(rows)


def style_schedule_dataframe(df: pd.DataFrame):
    employee_cols = [c for c in df.columns if c not in ["יום בחודש", "יום בשבוע", "מחלקה", "מיון", "שישי בוקר"]]
    styles = pd.DataFrame("", index=df.index, columns=df.columns)

    for idx, row in df.iterrows():
        is_weekend = row["יום בשבוע"] in ["שישי", "שבת"]
        for col in df.columns:
            if is_weekend and col in ["יום בחודש", "יום בשבוע", "מחלקה", "מיון", "שישי בוקר"]:
                styles.loc[idx, col] = "background-color: #fff200; font-weight: bold;"
            elif col in employee_cols and row[col] == "V":
                styles.loc[idx, col] = "background-color: #b7b7b7; color: #000000; font-weight: bold; text-align: center;"
            elif idx % 2 == 1:
                styles.loc[idx, col] = "background-color: #fafafa;"

    return df.style.apply(lambda _: styles, axis=None)


def _excel_addr(col_idx: int, row_idx: int, absolute: bool = True) -> str:
    col = get_column_letter(col_idx)
    if absolute:
        return f"${col}${row_idx}"
    return f"{col}{row_idx}"


def _excel_range(c1: int, r1: int, c2: int, r2: int) -> str:
    return f"${get_column_letter(c1)}${r1}:${get_column_letter(c2)}${r2}"


def dataframe_to_xlsx_bytes(parsed_df: pd.DataFrame, schedule_df: pd.DataFrame, year: int, month: int) -> bytes:
    """
    Export a dynamic workbook similar to the uploaded duty-planning file.

    Structure:
    A:E = date/week/day/manual assignment columns.
    F:... = employee availability columns, gray + V for blocked/vacation days.
    After employees = dynamic calculation block:
      שם תורן | מצוי | רצוי | סופש | רצוי סופש | חמישי/שישי | רצוי חמישי
    Bottom = editable target table used by formulas.
    """
    buffer = BytesIO()
    employee_names = parsed_df["שם עובד/ת"].tolist() if not parsed_df.empty else []
    n_workers = len(employee_names)
    current_last_day = calendar.monthrange(year, month)[1]

    # Row model: header row 1, two previous month rows 2-3, current month rows 4..current_end_row.
    current_start_row = 4
    current_end_row = current_start_row + current_last_day - 1
    data_end_row = current_end_row

    first_employee_col = 6  # F
    last_employee_col = first_employee_col + n_workers - 1 if n_workers else first_employee_col - 1

    calc_start_col = first_employee_col + n_workers + 1
    calc_name_col = calc_start_col
    calc_found_col = calc_start_col + 1
    calc_target_col = calc_start_col + 2
    calc_weekend_col = calc_start_col + 3
    calc_weekend_target_col = calc_start_col + 4
    calc_thu_fri_col = calc_start_col + 5
    calc_thu_fri_target_col = calc_start_col + 6

    calc_header_row = 6
    calc_first_worker_row = 7
    calc_last_worker_row = calc_first_worker_row + n_workers - 1
    calc_total_row = calc_last_worker_row + 1

    # Bottom target table: names and manual desired counts.
    target_names_row = data_end_row + 4
    target_duties_row = target_names_row + 1
    target_weekends_row = target_names_row + 2
    target_thu_fri_row = target_names_row + 3

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        schedule_df.to_excel(writer, sheet_name="תורנויות", index=False)
        parsed_df.to_excel(writer, sheet_name="סיכום קלט", index=False)

        wb = writer.book
        ws = wb["תורנויות"]
        ws.sheet_view.rightToLeft = True

        # Basic fills and styles.
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        weekend_fill = PatternFill("solid", fgColor="FFF200")
        blocked_fill = PatternFill("solid", fgColor="B7B7B7")
        calc_fill = PatternFill("solid", fgColor="EAF4FF")
        target_fill = PatternFill("solid", fgColor="F3F4F6")
        white_fill = PatternFill("solid", fgColor="FFFFFF")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header style.
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Main table body styles.
        max_col = ws.max_column
        for row_idx in range(2, data_end_row + 1):
            weekday = ws.cell(row=row_idx, column=2).value
            is_weekend = weekday in ["שישי", "שבת"]

            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                if row_idx % 2 == 1:
                    cell.fill = white_fill

                # Yellow background for weekends in A:E inclusive.
                if is_weekend and col_idx <= 5:
                    cell.fill = weekend_fill
                    cell.font = Font(bold=True)

                # Gray + V in employee columns.
                if first_employee_col <= col_idx <= last_employee_col and cell.value == "V":
                    cell.fill = blocked_fill
                    cell.font = Font(bold=True, color="000000")

        # Data validation for manual assignment cells C:E.
        if n_workers:
            names_formula_range = _excel_range(first_employee_col, target_names_row, last_employee_col, target_names_row)
            for col_idx in [3, 4, 5]:
                # Use bottom names row as source. Some Excel versions dislike quoted sheet names in validation
                # inside same sheet, so local absolute range is used.
                ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                dv = {
                    "type": "list",
                    "formula1": f"={names_formula_range}",
                    "allow_blank": True,
                }
                # openpyxl data validation direct API is intentionally avoided here for Streamlit runtime stability.
                # Users can type names manually; formulas are based on exact text.

        # Calculation block headers.
        calc_headers = ["שם תורן", "מצוי", "רצוי", "סופש", "רצוי סופש", "חמישי/שישי", "רצוי חמישי"]
        for offset, header in enumerate(calc_headers):
            cell = ws.cell(row=calc_header_row, column=calc_start_col + offset)
            cell.value = header
            cell.fill = calc_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Calculation rows and formulas.
        assignment_range = _excel_range(3, current_start_row, 4, current_end_row)  # C:D
        target_names_range = _excel_range(first_employee_col, target_names_row, last_employee_col, target_names_row) if n_workers else ""
        target_duties_range = _excel_range(first_employee_col, target_duties_row, last_employee_col, target_duties_row) if n_workers else ""
        target_weekends_range = _excel_range(first_employee_col, target_weekends_row, last_employee_col, target_weekends_row) if n_workers else ""
        target_thu_fri_range = _excel_range(first_employee_col, target_thu_fri_row, last_employee_col, target_thu_fri_row) if n_workers else ""

        weekend_terms = []
        thu_fri_terms = []
        for row_idx in range(current_start_row, current_end_row + 1):
            weekday = ws.cell(row=row_idx, column=2).value
            if weekday in ["שישי", "שבת"]:
                weekend_terms.append(f"COUNTIF($C${row_idx}:$D${row_idx},{get_column_letter(calc_name_col)}{{row}})")
            if weekday == "חמישי":
                thu_fri_terms.append(f"COUNTIF($C${row_idx}:$D${row_idx},{get_column_letter(calc_name_col)}{{row}})")
            if weekday == "שישי":
                thu_fri_terms.append(f"COUNTIF($E${row_idx},{get_column_letter(calc_name_col)}{{row}})")

        for i, worker in enumerate(employee_names):
            row_idx = calc_first_worker_row + i
            name_cell = ws.cell(row=row_idx, column=calc_name_col)
            name_cell.value = worker
            name_cell.font = Font(bold=True)
            name_cell.alignment = Alignment(horizontal="center", vertical="center")
            name_cell.border = border

            ws.cell(row=row_idx, column=calc_found_col).value = f'=COUNTIF({assignment_range},{get_column_letter(calc_name_col)}{row_idx})'
            ws.cell(row=row_idx, column=calc_target_col).value = f'=IFERROR(INDEX({target_duties_range},1,MATCH({get_column_letter(calc_name_col)}{row_idx},{target_names_range},0)),"")'
            ws.cell(row=row_idx, column=calc_weekend_col).value = "=" + ("+".join(term.format(row=row_idx) for term in weekend_terms) if weekend_terms else "0")
            ws.cell(row=row_idx, column=calc_weekend_target_col).value = f'=IFERROR(INDEX({target_weekends_range},1,MATCH({get_column_letter(calc_name_col)}{row_idx},{target_names_range},0)),"")'
            ws.cell(row=row_idx, column=calc_thu_fri_col).value = "=" + ("+".join(term.format(row=row_idx) for term in thu_fri_terms) if thu_fri_terms else "0")
            ws.cell(row=row_idx, column=calc_thu_fri_target_col).value = f'=IFERROR(INDEX({target_thu_fri_range},1,MATCH({get_column_letter(calc_name_col)}{row_idx},{target_names_range},0)),"")'

            for col_idx in range(calc_start_col, calc_thu_fri_target_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        # Totals row.
        if n_workers:
            ws.cell(row=calc_total_row, column=calc_name_col).value = "סהכ"
            ws.cell(row=calc_total_row, column=calc_name_col).font = Font(bold=True)
            for col_idx in range(calc_found_col, calc_thu_fri_target_col + 1):
                col_letter = get_column_letter(col_idx)
                ws.cell(row=calc_total_row, column=col_idx).value = f"=SUM({col_letter}{calc_first_worker_row}:{col_letter}{calc_last_worker_row})"
                ws.cell(row=calc_total_row, column=col_idx).font = Font(bold=True)

            for col_idx in range(calc_start_col, calc_thu_fri_target_col + 1):
                cell = ws.cell(row=calc_total_row, column=col_idx)
                cell.fill = calc_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        # Bottom target table.
        if n_workers:
            for i, worker in enumerate(employee_names):
                col_idx = first_employee_col + i
                ws.cell(row=target_names_row, column=col_idx).value = worker
                ws.cell(row=target_names_row, column=col_idx).font = Font(bold=True)
                ws.cell(row=target_names_row, column=col_idx).fill = target_fill
                ws.cell(row=target_names_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=target_names_row, column=col_idx).border = border

                # Leave these values editable by default. Put 0 as a clear placeholder.
                for target_row in [target_duties_row, target_weekends_row, target_thu_fri_row]:
                    ws.cell(row=target_row, column=col_idx).value = 0
                    ws.cell(row=target_row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=target_row, column=col_idx).border = border

            labels = {
                target_duties_row: "תורנויות",
                target_weekends_row: "סופשים",
                target_thu_fri_row: "חמישי/שישי",
            }
            for row_idx, label in labels.items():
                ws.cell(row=row_idx, column=5).value = label
                ws.cell(row=row_idx, column=5).font = Font(bold=True)
                ws.cell(row=row_idx, column=5).fill = target_fill
                ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=row_idx, column=5).border = border

        # Notes.
        ws.cell(row=target_thu_fri_row + 4, column=2).value = "ב+ה חותמים"
        ws.cell(row=target_thu_fri_row + 5, column=2).value = "ג+ד+ו לא חותמים"

        # Freeze panes and widths.
        ws.freeze_panes = "F2"

        widths = {
            1: 10,
            2: 16,
            3: 15,
            4: 15,
            5: 15,
        }

        # Employees.
        for col_idx in range(first_employee_col, last_employee_col + 1):
            widths[col_idx] = 12

        # Calculations.
        for col_idx in range(calc_start_col, calc_thu_fri_target_col + 1):
            widths[col_idx] = 14

        for col_idx in range(1, calc_thu_fri_target_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 12)

        # Summary input sheet formatting.
        summary_ws = wb["סיכום קלט"]
        summary_ws.sheet_view.rightToLeft = True
        summary_ws.freeze_panes = "A2"
        for cell in summary_ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row in summary_ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
        for col in summary_ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            summary_ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 32)

    return buffer.getvalue()


def render_general_table_coding():
    st.header("קידוד לטבלה כללית")
    st.caption("הדבק כאן את הפלטים מכל התורנים. הכלי יוצר קובץ Excel מוכן לתיאום סופי עם חסימות, סימוני V, צביעות ונוסחאות דינמיות.")

    st.caption("הדבק כאן ברצף את הפלטים מכל העובדים/ות כפי שהתקבלו מכלי תכנון התורנויות.")

    raw_text = st.text_area(
        "הדבק פלטים של כל התורנים",
        height=300,
        placeholder='עובד/ת מס׳ 1\nחסימות- 1,2,3\nחופשים- 5\n\nעובד/ת מס׳ 2\nחסימות- 7,8\nחופשים- 12',
    )

    default_y, default_m = default_next_month()
    col_year, col_month = st.columns(2)
    with col_year:
        selected_year = st.number_input("שנה", min_value=2024, max_value=2100, value=default_y, step=1)
    with col_month:
        selected_month = st.number_input("חודש", min_value=1, max_value=12, value=default_m, step=1)

    st.info(
        "הקובץ יכלול את שני הימים האחרונים של החודש הקודם, את כל ימי החודש הנבחר, "
        "עמודות מחלקה/מיון/שישי בוקר למילוי ידני, עמודות תורנים, ונוסחאות חישוב אוטומטיות אחרי עמודות התורנים."
    )

    if st.button("צור טבלת תורנויות", type="primary"):
        if not raw_text.strip():
            st.warning("יש להדביק לפחות פלט אחד.")
            return

        parsed_df = parse_worker_blocks(raw_text)
        if parsed_df.empty:
            st.error("לא זוהו תורנים. בדוק שהמבנה הוא: שם, חסימות, חופשים.")
            return

        schedule_df = build_schedule_layout_table(parsed_df, int(selected_year), int(selected_month))

        st.success(f"זוהו {len(parsed_df)} תורנים עבור {month_title(int(selected_year), int(selected_month))}.")

        st.subheader("סיכום תורנים")
        display_df = parsed_df.copy()
        display_df["שם עובד/ת"] = [f"עובד/ת מס׳ {i}" for i in range(1, len(display_df) + 1)]
        st.dataframe(display_df, hide_index=True, use_container_width=True)
        st.caption("שמות העובדים/ות מוסתרים בתצוגה זו. בקובץ ה-Excel עצמו השמות נשמרים לצורך נוסחאות ושיבוץ.")

        st.subheader("תצוגה מקדימה")
        st.caption("בתצוגה: V בתאים חסומים/חופשים, אפור בעמודות תורנים, צהוב בסופי שבוע בעמודות A:E.")
        st.dataframe(style_schedule_dataframe(schedule_df), hide_index=True, use_container_width=True)

        xlsx_bytes = dataframe_to_xlsx_bytes(parsed_df, schedule_df, int(selected_year), int(selected_month))
        st.download_button(
            "הורד קובץ Excel לתיאום תורנויות",
            data=xlsx_bytes,
            file_name=f"tornuyot_{int(selected_year)}_{int(selected_month):02d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )



def _clean_worker_name(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("*", "").strip()
    text = re.sub(r"\s+", " ", text)
    # Values in the assignment sheet can be like "ב-נורית" or "ה-גל".
    # The calendar event belongs to the worker after the hyphen.
    if "-" in text:
        text = text.split("-")[-1].strip()
    if "–" in text:
        text = text.split("–")[-1].strip()
    return text


def _raw_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _find_schedule_sheet(wb):
    if "תורנויות" in wb.sheetnames:
        return wb["תורנויות"]
    return wb[wb.sheetnames[0]]


def _detect_workers_from_sheet(ws) -> List[str]:
    workers = []
    # In the generated/final file, worker availability columns start at F.
    # Stop at first empty cell or when calculation headers begin.
    for col in range(6, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        name = _clean_worker_name(value)
        if not name:
            break
        if name in {"מצוי", "רצוי", "סופש", "רצוי סופש", "חמישי/שישי", "שישי/חמישי", "רצוי חמישי", "שם תורן"}:
            break
        workers.append(name)

    # Fallback: detect from assignment columns C:E.
    if not workers:
        found = set()
        for row in range(2, ws.max_row + 1):
            day_value = ws.cell(row=row, column=1).value
            if not isinstance(day_value, int):
                continue
            for col in [3, 4, 5]:
                name = _clean_worker_name(ws.cell(row=row, column=col).value)
                if name:
                    found.add(name)
        workers = sorted(found)

    # Preserve order, remove duplicates.
    ordered = []
    seen = set()
    for worker in workers:
        if worker and worker not in seen:
            ordered.append(worker)
            seen.add(worker)
    return ordered


def _extract_month_rows(ws, year: int, month: int) -> List[dict]:
    last_day = calendar.monthrange(year, month)[1]
    first_month_row = None

    for row in range(2, ws.max_row + 1):
        day_value = ws.cell(row=row, column=1).value
        weekday_value = ws.cell(row=row, column=2).value
        if day_value == 1 and weekday_value:
            first_month_row = row
            break

    if first_month_row is None:
        return []

    rows = []
    for row in range(first_month_row, min(ws.max_row, first_month_row + last_day + 5) + 1):
        day_value = ws.cell(row=row, column=1).value
        weekday_value = ws.cell(row=row, column=2).value
        if not isinstance(day_value, int):
            continue
        if not (1 <= day_value <= last_day):
            continue

        rows.append({
            "excel_row": row,
            "date": date(year, month, day_value),
            "day_in_month": day_value,
            "weekday": str(weekday_value or ""),
            "מחלקה": _raw_text(ws.cell(row=row, column=3).value),
            "מיון": _raw_text(ws.cell(row=row, column=4).value),
            "שישי בוקר": _raw_text(ws.cell(row=row, column=5).value),
        })

        if day_value == last_day:
            break

    return rows


def parse_final_schedule_excel(uploaded_file, year: int, month: int) -> Tuple[List[str], List[dict]]:
    wb = load_workbook(uploaded_file, data_only=True)
    ws = _find_schedule_sheet(wb)
    workers = _detect_workers_from_sheet(ws)
    rows = _extract_month_rows(ws, year, month)
    return workers, rows



def extract_worker_events(schedule_rows: List[dict], worker_name: str, times_config: dict) -> List[dict]:
    events = []

    shift_definitions = [
        {
            "column": "מחלקה",
            "title": "תורנות מחלקה - פנימית ד׳",
            "start_time": times_config["department_start"],
            "end_time": times_config["department_end"],
            "ends_next_day": True,
        },
        {
            "column": "מיון",
            "title": "תורנות מיון - פנימית ד׳",
            "start_time": times_config["er_start"],
            "end_time": times_config["er_end"],
            "ends_next_day": True,
        },
        {
            "column": "שישי בוקר",
            "title": "שישי בוקר - פנימית ד׳",
            "start_time": times_config["friday_morning_start"],
            "end_time": times_config["friday_morning_end"],
            "ends_next_day": False,
        },
    ]

    selected = _clean_worker_name(worker_name)
    auto_friday_keys = set()

    for row in schedule_rows:
        for shift in shift_definitions:
            raw_value = row.get(shift["column"], "")
            assigned_worker = _clean_worker_name(raw_value)
            if not assigned_worker or assigned_worker != selected:
                continue

            start_dt = datetime.combine(row["date"], shift["start_time"])
            end_date = row["date"] + timedelta(days=1) if shift["ends_next_day"] else row["date"]
            end_dt = datetime.combine(end_date, shift["end_time"])

            events.append({
                "date": row["date"],
                "weekday": row["weekday"],
                "shift_type": shift["column"],
                "title": shift["title"],
                "start_dt": start_dt,
                "end_dt": end_dt,
                "raw_assignment": raw_value,
                "source": "שיבוץ ישיר",
            })

            # Business rule:
            # Department Saturday duty also means the Saturday duty resident works
            # Friday morning in the department, even if the Friday morning column
            # does not explicitly contain the worker name.
            if shift["column"] == "מחלקה" and row["weekday"] == "שבת":
                friday_date = row["date"] - timedelta(days=1)
                friday_start = datetime.combine(friday_date, times_config["friday_morning_start"])
                friday_end = datetime.combine(friday_date, times_config["friday_morning_end"])
                key = (selected, friday_start, friday_end, "שישי בוקר במחלקה")
                if key not in auto_friday_keys:
                    auto_friday_keys.add(key)
                    events.append({
                        "date": friday_date,
                        "weekday": "שישי",
                        "shift_type": "שישי בוקר",
                        "title": "שישי בוקר במחלקה - פנימית ד׳",
                        "start_dt": friday_start,
                        "end_dt": friday_end,
                        "raw_assignment": f"אוטומטי מתורנות מחלקה בשבת ({row['date'].strftime('%d/%m/%Y')})",
                        "source": "אוטומטי מתורנות שבת",
                    })

    deduped = []
    seen = set()
    for event in events:
        key = (event["title"], event["start_dt"], event["end_dt"], event["shift_type"])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(event)

    deduped.sort(key=lambda item: item["start_dt"])
    return deduped


def _ics_escape(text: str) -> str:
    text = str(text or "")
    text = text.replace("\\", "\\\\")
    text = text.replace(";", "\\;")
    text = text.replace(",", "\\,")
    text = text.replace("\n", "\\n")
    return text


def _ics_dt(dt: datetime) -> str:
    return dt.strftime("%Y%m%dT%H%M%S")


def build_ics(events: List[dict], worker_name: str, calendar_name: str = "תורנויות פנימית ד׳") -> str:
    now = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//MedStaff//PnimitD Scheduling Tool//HE",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:{_ics_escape(calendar_name)}",
        "X-WR-TIMEZONE:Asia/Jerusalem",
    ]

    safe_worker = _clean_worker_name(worker_name)

    for idx, event in enumerate(events, start=1):
        uid = f"pnimit-d-{safe_worker}-{event['start_dt'].strftime('%Y%m%d%H%M')}-{idx}@medstaff.local"
        description = (
            f"תורן/ית: {safe_worker}\\n"
            f"סוג תורנות: {event['shift_type']}\\n"
            f"יום: {event['weekday']}\\n"
            f"מקור: {event.get('source', 'שיבוץ ישיר')}\\n"
            f"שיבוץ מקורי בקובץ: {event['raw_assignment']}"
        )

        lines.extend([
            "BEGIN:VEVENT",
            f"UID:{_ics_escape(uid)}",
            f"DTSTAMP:{now}",
            f"DTSTART;TZID=Asia/Jerusalem:{_ics_dt(event['start_dt'])}",
            f"DTEND;TZID=Asia/Jerusalem:{_ics_dt(event['end_dt'])}",
            f"SUMMARY:{_ics_escape(event['title'])}",
            f"DESCRIPTION:{_ics_escape(description)}",
            "LOCATION:פנימית ד׳",
            "BEGIN:VALARM",
            "TRIGGER:-PT12H",
            "ACTION:DISPLAY",
            f"DESCRIPTION:{_ics_escape(event['title'])}",
            "END:VALARM",
            "END:VEVENT",
        ])

    lines.append("END:VCALENDAR")
    return "\r\n".join(lines) + "\r\n"


def render_calendar_save():
    st.header("שמירה ביומן")
    st.caption("העלה את קובץ ה-Excel הסופי לאחר חלוקת התורנויות, בחר תורן, והורד קובץ ICS לייבוא ליומן.")

    st.info(
        "קובץ ICS הוא פורמט יומן אחיד שנתמך ב-Google Calendar, Outlook ו-Apple Calendar. "
        "אין צורך בחיבור ישיר ליומן או בהרשאות OAuth."
    )

    uploaded_file = st.file_uploader(
        "העלה קובץ תורנויות סופי",
        type=["xlsx"],
        help="הקובץ צריך לכלול את עמודות: יום בחודש, יום בשבוע, מחלקה, מיון, שישי בוקר.",
    )

    default_y, default_m = default_current_month()
    col_year, col_month = st.columns(2)
    with col_year:
        selected_year = st.number_input("שנה", min_value=2024, max_value=2100, value=default_y, step=1, key="ics_year")
    with col_month:
        selected_month = st.number_input("חודש", min_value=1, max_value=12, value=default_m, step=1, key="ics_month")

    st.subheader("שעות תורנות")
    time_col1, time_col2, time_col3 = st.columns(3)
    with time_col1:
        department_start = st.time_input("מחלקה - התחלה", value=time(8, 0))
        department_end = st.time_input("מחלקה - סיום למחרת", value=time(10, 0))
    with time_col2:
        er_start = st.time_input("מיון - התחלה", value=time(16, 0))
        er_end = st.time_input("מיון - סיום למחרת", value=time(8, 0))
    with time_col3:
        friday_morning_start = st.time_input("שישי בוקר - התחלה", value=time(8, 0))
        friday_morning_end = st.time_input("שישי בוקר - סיום", value=time(14, 0))

    if not uploaded_file:
        st.caption("לאחר העלאת קובץ, תופיע בחירת תורן ורשימת התורנויות שלו.")
        return

    try:
        workers, schedule_rows = parse_final_schedule_excel(uploaded_file, int(selected_year), int(selected_month))
    except Exception as e:
        st.error(f"שגיאה בקריאת הקובץ: {e}")
        return

    if not schedule_rows:
        st.error("לא נמצאו שורות חודש תקינות בקובץ. ודא שהעמודה הראשונה כוללת את ימי החודש ושיש שורה עבור יום 1.")
        return

    if not workers:
        st.warning("לא זוהו שמות תורנים מהקובץ. ניתן עדיין להקליד שם ידנית.")
        worker_name = st.text_input("שם תורן", value="")
    else:
        worker_name = st.selectbox("בחר תורן", workers)

    if not worker_name:
        return

    times_config = {
        "department_start": department_start,
        "department_end": department_end,
        "er_start": er_start,
        "er_end": er_end,
        "friday_morning_start": friday_morning_start,
        "friday_morning_end": friday_morning_end,
    }

    events = extract_worker_events(schedule_rows, worker_name, times_config)

    st.subheader("תורנויות שיזוהו ליומן")
    if not events:
        st.warning("לא נמצאו תורנויות עבור התורן שנבחר.")
        return

    preview_df = pd.DataFrame([
        {
            "תאריך": event["date"].strftime("%d/%m/%Y"),
            "יום": event["weekday"],
            "סוג": event["shift_type"],
            "התחלה": event["start_dt"].strftime("%d/%m/%Y %H:%M"),
            "סיום": event["end_dt"].strftime("%d/%m/%Y %H:%M"),
            "מקור": event.get("source", "שיבוץ ישיר"),
            "שיבוץ בקובץ": event["raw_assignment"],
        }
        for event in events
    ])

    # Reverse technical order to keep visual RTL.
    preview_order = ["שיבוץ בקובץ", "מקור", "סיום", "התחלה", "סוג", "יום", "תאריך"]
    st.dataframe(preview_df[preview_order], hide_index=True, use_container_width=True)

    ics_content = build_ics(events, worker_name)
    safe_name = re.sub(r"[^A-Za-z0-9א-ת_-]+", "_", _clean_worker_name(worker_name)).strip("_") or "worker"

    st.download_button(
        "הורד קובץ יומן ICS",
        data=ics_content.encode("utf-8"),
        file_name=f"tornuyot_{safe_name}_{int(selected_year)}_{int(selected_month):02d}.ics",
        mime="text/calendar",
        use_container_width=True,
    )

    st.caption("לאחר ההורדה, פתח את הקובץ או ייבא אותו ליומן הרצוי. Google, Outlook ו-Apple Calendar תומכים ב-ICS.")


def render_sidebar_navigation() -> str:
    st.sidebar.title("מע׳ לתכנון תורנויות- פנימית ד׳")
    st.sidebar.markdown("### כלי עזר")
    st.sidebar.markdown("---")

    return st.sidebar.radio(
        "בחר כלי",
        [
            "תכנון תורנויות",
            "קידוד לטבלה כללית",
            "שמירה ביומן",
        ],
        index=0,
        label_visibility="collapsed",
    )


def main():
    inject_global_rtl_css()
    selected_tool = render_sidebar_navigation()


    if selected_tool == "תכנון תורנויות":
        render_shift_planning()
    elif selected_tool == "קידוד לטבלה כללית":
        render_general_table_coding()
    elif selected_tool == "שמירה ביומן":
        render_calendar_save()


if __name__ == "__main__":
    main()
